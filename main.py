from pgmagick import Image, DrawableRectangle, DrawableText, Geometry, Color, DrawableList, DrawableGravity, GravityType, CompositeOperator as co
import requests
import os.path
from os import path
import openpyxl
import argparse

parser = argparse.ArgumentParser(prog = 'Skyrim Item Generator')
parser.add_argument('-f', '--file')
args = parser.parse_args()

itemData = []

# load excel with its path
workbook = openpyxl.load_workbook("./input/" + args.file + ".xlsx")
sheet = workbook.active
for i in range(4, sheet.max_row+1):
	singleItemData = {}
	for j in range(1, sheet.max_column+1):
		object_type = sheet.cell(row=3, column=j)
		cell_obj = sheet.cell(row=i, column=j)
		tempVal = cell_obj.value
		if tempVal == None:
			tempVal = ""
		if type(tempVal) != type("test"):
			tempVal = int(tempVal)
		singleItemData[object_type.value] = tempVal
	itemData.append(singleItemData)

for item in itemData:
	if item["itemName"] != "":
		base = Image('.\Item_Blank.png')
		base.font(".\Adobe Garamond Pro Bold.ttf")

		# Handle item image
		tempItemName = item["itemName"].replace('\n', '_').replace(' ', '_')
		if not path.exists('./temp/' + str(tempItemName) + '.jpg') and item["itemPictureURL"] != "":
			itemPictureData = requests.get(item["itemPictureURL"]).content
			with open('./temp/' + str(tempItemName) + '.jpg', 'wb') as handler:
				handler.write(itemPictureData)
		if item["itemPictureURL"] != "":
			tempIM = Image(Geometry(225,225,0,0), Color(0,0,255,1))
			itemPicture = Image('./temp/' + str(tempItemName) + '.jpg')
			itemPicture.resize('225x225')
			itemPicture.extent(Geometry(225,225,0,0), Color(0,0,0,1), GravityType.CenterGravity)
			base.composite(itemPicture, 150, 170, co.LightenCompositeOp)

		# Add loot type symbol
		typeImage = Image('.\Item_TreasureIcon_' + item["lootType"] + '.png')
		base.composite(typeImage, 0, 0, co.OverCompositeOp)
		# Take care of numbered quest items
		if item["lootType"] == "S":
			base.fontPointsize(55)
			base.annotate(str(item["specialLootNumber"]), Geometry(1,1,80,80), GravityType.CenterGravity, -45)


		# Add item type symbol
		itemTypeImage = Image('.\Item_Type_' + item["itemType"] + '.png')
		base.composite(itemTypeImage, 0, 0, co.OverCompositeOp)

		# Add item name
		base.fontPointsize(40)
		base.fillColor("White")
		base.annotate(item["itemName"], Geometry(1,1,260,90), GravityType.CenterGravity)
		base.fillColor("Black")

		# Add enchantment/upgrade costs
		enchantmentCostImage = Image('.\Item_EnchantmentCost.png')
		upgradeCostImage = Image('.\Item_EnchantmentUpgradeCost.png')
		if int(item["upgradeCost"]) > 0:
			base.composite(upgradeCostImage, 0, 0, co.OverCompositeOp)
			base.fontPointsize(55)
			base.annotate(str(item["enchantmentCost"]), Geometry(1,1,420,686), GravityType.CenterGravity)
			base.fontPointsize(55)
			base.annotate(str(item["upgradeCost"]), Geometry(1,1,460,715), GravityType.CenterGravity)
		elif int(item["enchantmentCost"]) > 0:
			base.composite(enchantmentCostImage, 0, 0, co.OverCompositeOp)
			base.fontPointsize(65)
			base.annotate(str(item["enchantmentCost"]), Geometry(1,1,432,700), GravityType.CenterGravity)


		# Add price symbol
		if int(item["cost"]) > 0:
			priceImage = Image('.\Item_Price.png')
			base.composite(priceImage, 0, 0, co.OverCompositeOp)

			# Write item price
			base.fontPointsize(85)
			base.annotate(str(item["cost"]), Geometry(1,1,80,710), GravityType.CenterGravity)



		# Add skills
		if item["skill1Name"] != "":
			skills = "Single"
			if item["skill2Name"] != "":
				skills = "Double"
			if item["skill3Name"] != "":
				skills = "Triple"
			skillFrameImage = Image('.\Item_Skill_'+skills+'.png')
			base.composite(skillFrameImage, 0, 0, co.OverCompositeOp)

			# Skill 1 stuff -----------------------------------------------------------------------
			skillYVal = 426
			# Boost bounding box
			im = Image(Geometry(81, 53), Color(item["skill1BoostType"]))
			skill1BoostRect = DrawableRectangle(0, 0, 0, 0)
			im.draw(skill1BoostRect)
			base.composite(im, 30, skillYVal - 26, co.OverCompositeOp)
			# Boost text
			base.fontPointsize(25)
			base.annotate(str(item["skill1Boost"]), Geometry(1,1,71,skillYVal), GravityType.CenterGravity)
			# Cost symbol
			skill1CostSymbolImage = Image('.\Symbol_'+item["skill1CostSymbol"]+'.png')
			skill1CostSymbolImage.resize('40x40')
			base.composite(skill1CostSymbolImage, 165-20, skillYVal-20, co.OverCompositeOp)
			# Cost number
			base.fontPointsize(40)
			base.annotate(str(item["skill1Cost"]), Geometry(1,1,132,skillYVal), GravityType.CenterGravity)
			# Name text
			base.fontPointsize(25)
			base.annotate(str(item["skill1Name"]), Geometry(1,1,234,skillYVal), GravityType.CenterGravity)
			# Result symbol
			skill1ResultSymbolImage = Image('.\Symbol_'+item["skill1ResultSymbol"]+'.png')
			skill1ResultSymbolImage.resize('40x40')
			base.composite(skill1ResultSymbolImage, 324-20, skillYVal-20, co.OverCompositeOp)
			# Result number
			base.fontPointsize(40)
			base.annotate(str(item["skill1Result"]), Geometry(1,1,297,skillYVal), GravityType.CenterGravity)
			# Damage symbol 1
			if item["skill1ResultDamageSymbol1"] != "":
				skill1ResultDamageSymbol1Image = Image('.\Symbol_'+item["skill1ResultDamageSymbol1"]+'.png')
				skill1ResultDamageSymbol1Image.resize('40x40')
				base.composite(skill1ResultDamageSymbol1Image, 395-20, skillYVal - 20, co.OverCompositeOp)
			# Damage symbol 2
			if item["skill1ResultDamageSymbol2"] != "":
				skill1ResultDamageSymbol2Image = Image('.\Symbol_'+item["skill1ResultDamageSymbol2"]+'.png')
				skill1ResultDamageSymbol2Image.resize('40x40')
				base.composite(skill1ResultDamageSymbol2Image, 431-20, skillYVal - 20, co.OverCompositeOp)
			# Damage symbol 3
			if item["skill1ResultDamageSymbol3"] != "":
				skill1ResultDamageSymbol3Image = Image('.\Symbol_'+item["skill1ResultDamageSymbol3"]+'.png')
				skill1ResultDamageSymbol3Image.resize('40x40')
				base.composite(skill1ResultDamageSymbol3Image, 468-20, skillYVal - 20, co.OverCompositeOp)
			# Damage number text
			base.fontPointsize(40)
			base.annotate(str(item["skill1ResultDamage"]), Geometry(1,1,367,skillYVal), GravityType.CenterGravity)


			# Skill 2 stuff -----------------------------------------------------------------------
			if item["skill2Name"] != "":
				skillYVal = 488
				# Boost bounding box
				im = Image(Geometry(81, 53), Color(item["skill2BoostType"]))
				skill2BoostRect = DrawableRectangle(0, 0, 0, 0)
				im.draw(skill2BoostRect)
				base.composite(im, 30, skillYVal - 26, co.OverCompositeOp)
				# Boost text
				base.fontPointsize(25)
				base.annotate(str(item["skill2Boost"]), Geometry(1,1,71,skillYVal), GravityType.CenterGravity)
				# Cost symbol
				skill2CostSymbolImage = Image('.\Symbol_'+item["skill2CostSymbol"]+'.png')
				skill2CostSymbolImage.resize('40x40')
				base.composite(skill2CostSymbolImage, 165-20, skillYVal-20, co.OverCompositeOp)
				# Cost number
				base.fontPointsize(40)
				base.annotate(str(item["skill2Cost"]), Geometry(1,1,132,skillYVal), GravityType.CenterGravity)
				# Name text
				base.fontPointsize(25)
				base.annotate(str(item["skill2Name"]), Geometry(1,1,234,skillYVal), GravityType.CenterGravity)
				# Result symbol
				skill2ResultSymbolImage = Image('.\Symbol_'+item["skill2ResultSymbol"]+'.png')
				skill2ResultSymbolImage.resize('40x40')
				base.composite(skill2ResultSymbolImage, 324-20, skillYVal-20, co.OverCompositeOp)
				# Result number
				base.fontPointsize(40)
				base.annotate(str(item["skill2Result"]), Geometry(1,1,297,skillYVal), GravityType.CenterGravity)
				# Damage symbol 1
				if item["skill2ResultDamageSymbol1"] != "":
					skill2ResultDamageSymbol1Image = Image('.\Symbol_'+item["skill2ResultDamageSymbol1"]+'.png')
					skill2ResultDamageSymbol1Image.resize('40x40')
					base.composite(skill2ResultDamageSymbol1Image, 395-20, skillYVal - 20, co.OverCompositeOp)
				# Damage symbol 2
				if item["skill2ResultDamageSymbol2"] != "":
					skill2ResultDamageSymbol2Image = Image('.\Symbol_'+item["skill2ResultDamageSymbol2"]+'.png')
					skill2ResultDamageSymbol2Image.resize('40x40')
					base.composite(skill2ResultDamageSymbol2Image, 431-20, skillYVal - 20, co.OverCompositeOp)
				# Damage symbol 3
				if item["skill2ResultDamageSymbol3"] != "":
					skill2ResultDamageSymbol3Image = Image('.\Symbol_'+item["skill2ResultDamageSymbol3"]+'.png')
					skill2ResultDamageSymbol3Image.resize('40x40')
					base.composite(skill2ResultDamageSymbol3Image, 468-20, skillYVal - 20, co.OverCompositeOp)
				# Damage number text
				base.fontPointsize(40)
				base.annotate(str(item["skill2ResultDamage"]), Geometry(1,1,367,skillYVal), GravityType.CenterGravity)



			# Skill 3 stuff -----------------------------------------------------------------------
			if item["skill3Name"] != "":
				skillYVal = 550
				# Boost bounding box
				im = Image(Geometry(81, 53), Color(item["skill3BoostType"]))
				skill3BoostRect = DrawableRectangle(0, 0, 0, 0)
				im.draw(skill3BoostRect)
				base.composite(im, 30, skillYVal - 26, co.OverCompositeOp)
				# Boost text
				base.fontPointsize(25)
				base.annotate(str(item["skill3Boost"]), Geometry(1,1,71,skillYVal), GravityType.CenterGravity)
				# Cost symbol
				skill3CostSymbolImage = Image('.\Symbol_'+item["skill3CostSymbol"]+'.png')
				skill3CostSymbolImage.resize('40x40')
				base.composite(skill3CostSymbolImage, 165-20, skillYVal-20, co.OverCompositeOp)
				# Cost number
				base.fontPointsize(40)
				base.annotate(str(item["skill3Cost"]), Geometry(1,1,132,skillYVal), GravityType.CenterGravity)
				# Name text
				base.fontPointsize(25)
				base.annotate(str(item["skill3Name"]), Geometry(1,1,234,skillYVal), GravityType.CenterGravity)
				# Result symbol
				skill3ResultSymbolImage = Image('.\Symbol_'+item["skill3ResultSymbol"]+'.png')
				skill3ResultSymbolImage.resize('40x40')
				base.composite(skill3ResultSymbolImage, 324-20, skillYVal-20, co.OverCompositeOp)
				# Result number
				base.fontPointsize(40)
				base.annotate(str(item["skill3Result"]), Geometry(1,1,297,skillYVal), GravityType.CenterGravity)
				# Damage symbol 1
				if item["skill3ResultDamageSymbol1"] != "":
					skill3ResultDamageSymbol1Image = Image('.\Symbol_'+item["skill3ResultDamageSymbol1"]+'.png')
					skill3ResultDamageSymbol1Image.resize('40x40')
					base.composite(skill3ResultDamageSymbol1Image, 395-20, skillYVal - 20, co.OverCompositeOp)
				# Damage symbol 2
				if item["skill3ResultDamageSymbol2"] != "":
					skill3ResultDamageSymbol2Image = Image('.\Symbol_'+item["skill3ResultDamageSymbol2"]+'.png')
					skill3ResultDamageSymbol2Image.resize('40x40')
					base.composite(skill3ResultDamageSymbol2Image, 431-20, skillYVal - 20, co.OverCompositeOp)
				# Damage symbol 3
				if item["skill3ResultDamageSymbol3"] != "":
					skill3ResultDamageSymbol3Image = Image('.\Symbol_'+item["skill3ResultDamageSymbol3"]+'.png')
					skill3ResultDamageSymbol3Image.resize('40x40')
					base.composite(skill2ResultDamageSymbol3Image, 468-20, skillYVal - 20, co.OverCompositeOp)
				# Damage number text
				base.fontPointsize(40)
				base.annotate(str(item["skill3ResultDamage"]), Geometry(1,1,367,skillYVal), GravityType.CenterGravity)


		# Armor symbols
		heavyArmorImage = Image('.\Symbol_Heavy.png')
		lightArmorImage = Image('.\Symbol_Light.png')
		magicArmorImage = Image('.\Symbol_Magic.png')
		heavyArmorImage.resize('50x50')
		lightArmorImage.resize('50x50')
		magicArmorImage.resize('50x50')
		base.fillColor("White")
		if item["armorHeavy"] > 0:
			base.fontPointsize(45)
			base.annotate(str(item["armorHeavy"]), Geometry(1,1,150,700), GravityType.CenterGravity)
			base.composite(heavyArmorImage, 188-25, 700-25, co.OverCompositeOp)
		if item["armorLight"] > 0:
			base.fontPointsize(45)
			base.annotate(str(item["armorLight"]), Geometry(1,1,238,700), GravityType.CenterGravity)
			base.composite(lightArmorImage, 275-25, 700-25, co.OverCompositeOp)
		if item["armorMagic"] > 0:
			base.fontPointsize(45)
			base.annotate(str(item["armorMagic"]), Geometry(1,1,310,700), GravityType.CenterGravity)
			base.composite(magicArmorImage, 349-25, 700-25, co.OverCompositeOp)

		# Item description
		base.fontPointsize(27)
		if item["skill3Name"] != "":
			base.annotate(str(item["itemDescription"]), Geometry(1,1,250,628), GravityType.CenterGravity)
		elif item["skill2Name"] != "":
			base.annotate(str(item["itemDescription"]), Geometry(1,1,250,593), GravityType.CenterGravity)
		else:
			base.annotate(str(item["itemDescription"]), Geometry(1,1,250,550), GravityType.CenterGravity)
		base.fillColor("Black")

		# Write output
		base.write('./output/item_'+tempItemName+'.png')
